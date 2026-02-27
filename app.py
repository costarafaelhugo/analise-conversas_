import streamlit as st
import pandas as pd
import re
import json
import time
import csv
from io import StringIO, BytesIO
from typing import List, Dict
from datetime import datetime

# Versionamento semântico (MAJOR.MINOR.PATCH):
# MAJOR = mudança grande no modelo de análise ou comportamento (ex.: novo prompt de transbordo)
# MINOR = nova funcionalidade compatível (ex.: modo Analista de Categorias, novas colunas, nova taxonomia de motivos)
# PATCH = correções, ajustes de UI, documentação, scripts
# Histórico: 1.0 inicial → 1.x critérios/colunas/categorias → 2.0 prompt produção (transbordo) → 2.1 prompt com taxonomia causal de motivo_transbordo
APP_VERSION = "2.1.0"


# Configuração da página
st.set_page_config(
    page_title="Analista de Conversas - QA Chatbot",
    page_icon="🤖",
    layout="wide"
)

# Ocultar elementos do Streamlit Cloud (avatar do criador, footer e "created by")
st.markdown(
    """
    <style>
        /* Ocultar avatar do criador */
        ._profileContainer_gzau3_53,
        div[class*="profileContainer"],
        div[class*="profilePreview"],
        a[href*="share.streamlit.io/user"],
        img[alt="App Creator Avatar"],
        img[data-testid="appCreatorAvatar"] {
            display: none !important;
            visibility: hidden !important;
        }
        
        /* Ocultar qualquer elemento com profileContainer */
        div:has(._profileContainer_gzau3_53),
        div:has(div[class*="profileContainer"]) {
            display: none !important;
        }
        
        /* Ocultar elementos do Streamlit Cloud footer */
        footer,
        [data-testid="stFooter"],
        div[data-testid="stFooter"] {
            display: none !important;
            visibility: hidden !important;
        }
    </style>
    <script>
        // Ocultar elementos que contenham "created by" ou "hugo costa"
        function ocultarElementosCriador() {
            const textosParaOcultar = ['created by', 'Created by', 'CREATED BY', 'hugo costa', 'Hugo Costa', 'HUGO COSTA'];
            
            // Função recursiva para verificar todos os elementos
            function verificarElemento(elemento) {
                if (!elemento) return;
                
                const texto = elemento.textContent || elemento.innerText || '';
                const textoLower = texto.toLowerCase();
                
                // Verificar se o elemento ou seus filhos contêm os textos
                for (const textoProcurado of textosParaOcultar) {
                    if (textoLower.includes(textoProcurado.toLowerCase())) {
                        elemento.style.display = 'none';
                        elemento.style.visibility = 'hidden';
                        return;
                    }
                }
                
                // Verificar filhos
                if (elemento.children) {
                    for (const filho of elemento.children) {
                        verificarElemento(filho);
                    }
                }
            }
            
            // Executar quando a página carregar e periodicamente
            setTimeout(() => {
                document.querySelectorAll('*').forEach(el => {
                    verificarElemento(el);
                });
            }, 100);
            
            // Executar periodicamente para pegar elementos carregados dinamicamente
            setInterval(() => {
                document.querySelectorAll('*').forEach(el => {
                    verificarElemento(el);
                });
            }, 1000);
        }
        
        // Executar quando o DOM estiver pronto
        if (document.readyState === 'loading') {
            document.addEventListener('DOMContentLoaded', ocultarElementosCriador);
        } else {
            ocultarElementosCriador();
        }
    </script>
    """,
    unsafe_allow_html=True
)

# Título da aplicação
st.title("🤖 Analista de Conversas - QA Chatbot")
st.caption(f"Versão **{APP_VERSION}**")
st.markdown("---")

# Sidebar - Configurações
st.sidebar.header("⚙️ Configurações")
st.sidebar.caption(f"Versão **{APP_VERSION}**")

# Configurações para OpenAI API (obrigatório)
st.sidebar.markdown("---")
st.sidebar.subheader("🔑 Configurações OpenAI")

# Importar openai
try:
    import openai
except ImportError:
    st.sidebar.error("❌ Biblioteca openai não instalada. Execute: pip install openai")
    st.stop()

api_key = st.sidebar.text_input(
    "OpenAI API Key",
    value="",
    type="password",
    help="Insira sua chave da API do OpenAI"
)

model_name = st.sidebar.selectbox(
    "Modelo OpenAI",
    options=["gpt-4o-mini", "gpt-4o", "gpt-4-turbo", "gpt-3.5-turbo"],
    index=0,
    help="Selecione o modelo do OpenAI (gpt-4o-mini é mais rápido e econômico)"
)

# Configuração de delay entre requisições
delay_entre_requisicoes = st.sidebar.slider(
    "Delay entre requisições (segundos)",
    min_value=1,
    max_value=30,
    value=5,
    step=1,
    help="Aumente este valor se estiver recebendo erros de rate limit. Recomendado: 5-10 segundos para contas gratuitas, 3-5 para contas pagas."
)

st.sidebar.info("💡 **Dica**: Se receber erros de rate limit, aumente o delay entre requisições.")

# Configuração geral - Limite de conversas
st.sidebar.markdown("---")
st.sidebar.subheader("📊 Configurações de Processamento")

# Inicializar limite de conversas no session state
if 'limite_conversas' not in st.session_state:
    st.session_state['limite_conversas'] = None

# Verificar se há conversas carregadas no session state para mostrar o máximo
max_conversas = 1000  # Valor padrão
help_text = "Deixe vazio para analisar todas as conversas. Útil para testar com poucas conversas ou processar em lotes menores para evitar rate limits."

if 'conversas_carregadas_count' in st.session_state:
    max_conversas = max(1000, st.session_state['conversas_carregadas_count'])
    help_text = f"Deixe vazio para analisar todas as {st.session_state['conversas_carregadas_count']} conversas carregadas. Útil para testar com poucas conversas ou processar em lotes menores para evitar rate limits."

limite_conversas = st.sidebar.number_input(
    "Número máximo de conversas a analisar",
    min_value=1,
    max_value=max_conversas,
    value=None,
    step=1,
    help=help_text
)

# Salvar no session state
st.session_state['limite_conversas'] = limite_conversas if limite_conversas else None

# Mostrar informação sobre o limite
if limite_conversas:
    st.sidebar.info(f"📌 Limite ativo: **{limite_conversas} conversas**")
else:
    st.sidebar.info("📌 Sem limite: **Todas as conversas** serão analisadas")

# Função para extrair JSON do texto (para OpenAI)
def extract_json_from_text(text: str) -> Dict:
    """Extrai JSON do texto retornado pelo Gemini"""
    text = text.strip()
    
    # Tenta encontrar JSON entre ```json e ```
    json_block = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', text, re.DOTALL)
    if json_block:
        try:
            return json.loads(json_block.group(1))
        except json.JSONDecodeError:
            pass
    
    # Tenta encontrar JSON entre chaves (múltiplas linhas)
    json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', text, re.DOTALL)
    if json_match:
        try:
            return json.loads(json_match.group())
        except json.JSONDecodeError:
            pass
    
    # Tenta parsear todo o texto como JSON
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    
    return None

# Função para criar prompt do sistema
def criar_prompt_sistema(conversa: str) -> str:
    """Cria o prompt estruturado para análise da conversa via OpenAI"""
    prompt = f"""TAREFA:
Analisar a conversa entre CLIENTE e o sistema (WHIZZ + ATENDENTE BOT, avaliados como um único agente) e determinar:

1. Se houve NECESSIDADE REAL de transferência para atendimento humano
2. Qual foi o MOTIVO DO TRANSBORDO — mas SOMENTE se o transbordo foi efetivamente realizado

IMPORTANTE:
Este é um AGENTE DE ANÁLISE DE CONVERSAS.
Não é um agente de pós-vendas nem um agente operacional.
Avalie apenas o comportamento do sistema, sua aderência ao escopo e sua capacidade de conduzir corretamente a conversa.

---------------------------------------------------------------------

RETORNO (JSON EXATO – sem texto adicional)

Se NÃO houve transbordo efetivado:
{{
  "had_need_to_transfer": true ou false,
  "motivo_transbordo": null
}}

Se houve transbordo efetivado:
{{
  "had_need_to_transfer": true ou false,
  "motivo_transbordo": "categoria_padronizada"
}}

Nunca inventar motivo se não houve transbordo real.
Nunca inferir intenção.
Avaliar apenas eventos que ocorreram.

---------------------------------------------------------------------

REGRA CRÍTICA — DIFERENCIAÇÃO OBRIGATÓRIA

NÃO CONFUNDIR:

✔ necessidade de transbordo  
✔ transbordo efetivamente realizado  

O campo motivo_transbordo deve refletir SOMENTE:
→ transbordos que realmente aconteceram na conversa

Se o cliente pediu humano mas não foi transferido → motivo_transbordo = null

---------------------------------------------------------------------

PROCESSO OBRIGATÓRIO DE RACIOCÍNIO (NÃO EXIBIR)

PASSO 1 — verificar se houve transbordo real  
PASSO 2 — se houve, classificar o motivo  
PASSO 3 — avaliar se o transbordo foi causado por falha do sistema  
PASSO 4 — definir had_need_to_transfer  

---------------------------------------------------------------------

CLASSIFICAÇÃO CAUSAL DO TRANSBORDO

1. TRANSBORDO OPERACIONAL NECESSÁRIO
Limitação legítima do agente ou natureza do caso.
Não representa falha.

2. TRANSBORDO POR FALHA DE CONDUÇÃO
Erro cognitivo, decisão incorreta ou fricção evitável.

Somente o tipo 2 pode gerar had_need_to_transfer = true.

---------------------------------------------------------------------

TAXONOMIA OFICIAL — MOTIVO_TRANSBORDO

Usar EXATAMENTE um dos valores abaixo quando houver transbordo:

STATUS_PEDIDO_ATRASADO  
STATUS_PEDIDO_ENTREGUE_NAO_RECEBIDO  
ENDERECO_INCORRETO  
REEMBOLSO_OU_ESTORNO_ATRASADO  
DUVIDA_USO_CODIGO_RASTREIO  
STATUS_TICKET  
PEDIDO_DEVOLVIDO_LOGISTICA  

DETALHES_STATUS_TROCA_DEVOLUCAO  
PROBLEMA_VALE_TROCA  
EXCECAO_PRAZO_EXPIRADO  
PRAZO_ESTORNO  
PROBLEMA_CODIGO_POSTAGEM  

ALTERACAO_PEDIDO_EM_ANDAMENTO  
ALTERACAO_DADOS_CADASTRAIS  
ALTERACAO_FORMA_PAGAMENTO_OU_DEVOLUCAO  

SOLICITACAO_CANCELAMENTO  
DUVIDA_PEDIDO_CANCELADO  

FALHA_IA_LOOP_OU_ALUCINACAO  
PEDIDO_NAO_LOCALIZADO_PELA_IA  

DUVIDA_PRE_VENDA  
LOJA_FISICA  
PEDIDO_DIRETO_HUMANO  
ASSUNTO_FORA_DO_ESCOPO  
OUTROS

Se nenhum motivo for identificável → OUTROS

---------------------------------------------------------------------

ESCOPO DO AGENTE DE PÓS-VENDAS (COMPORTAMENTO CORRETO)

Considere comportamento correto quando o sistema:
- Informa status do pedido com identificador válido
- Informa rastreio apenas quando enviado
- Informa status de troca ou devolução
- Informa código de postagem
- Informa vale-troca apenas quando disponível
- Orienta processos de troca ou devolução
- Transborda corretamente quando necessário

---------------------------------------------------------------------

FORA DE ESCOPO DO AGENTE

- Cancelamentos
- Alterações de pedido
- Pedido atrasado (resolução ativa)
- Pré-venda
- Alterações cadastrais operacionais

Sistema deve se posicionar como pós-vendas.

---------------------------------------------------------------------

CASO PRIORITÁRIO (REGRA ABSOLUTA)

Se:
cliente não recebeu vale/estorno  
sistema informa prazo  
cliente insiste  
sistema entra em loop  

→ had_need_to_transfer = false

---------------------------------------------------------------------

CRITÉRIOS OBRIGATÓRIOS DE PONTO DE ATENÇÃO

1. Pedido de humano ignorado  
2. Falta de posicionamento como pós-vendas  
3. Loop de recepção  
4. Repetição sem avanço  
5. Tentativa de resolver fora do escopo  
6. Busca sem dados mínimos  
7. Solicitação incompleta de dados  
8. Transbordo causado por falha evitável  

---------------------------------------------------------------------

CRITÉRIOS DE NÃO ATENÇÃO

- Transbordo operacional correto
- Prazo informado corretamente
- Cliente abandona conversa
- Fora de escopo tratado corretamente
- Limitações informadas corretamente

---------------------------------------------------------------------

REGRAS FINAIS

- Avaliar causalidade do transbordo
- Avaliar apenas eventos reais
- Se não houve transbordo → motivo_transbordo = null
- Falha evitável → true
- Limitação legítima → false

---------------------------------------------------------------------

CONVERSA A SER ANALISADA:
{conversa}

IMPORTANTE:
Retorne APENAS o JSON final.
Sem explicações.
Sem texto adicional.
Sem comentários."""
    return prompt

# Função para analisar uma conversa via OpenAI API
def analisar_conversa_openai(conversa: str, modelo: str, api_key_openai: str = None) -> Dict:
    """Analisa uma conversa usando a API do OpenAI"""
    try:
        # Importar openai
        try:
            import openai
        except ImportError:
            return {
                "acao_necessaria": True,
                "tipo_falha": "Erro de dependência",
                "motivo_transbordo": "N/A",
                "descricao": "Erro: Biblioteca openai não está instalada. Execute: pip install openai",
                "sugestao_solucao": "Instalar biblioteca: pip install openai"
            }
        
        # Verificar API Key
        if not api_key_openai:
            return {
                "acao_necessaria": True,
                "tipo_falha": "Erro de configuração",
                "motivo_transbordo": "N/A",
                "descricao": "Erro: OpenAI API Key não foi configurada. Configure na barra lateral.",
                "sugestao_solucao": "Configurar OpenAI API Key na barra lateral da aplicação"
            }
        
        # Configurar cliente OpenAI
        client = openai.OpenAI(api_key=api_key_openai)
        
        # Verificar se a conversa não está vazia
        if not conversa or len(conversa.strip()) < 10:
            return {
                "acao_necessaria": False,
                "tipo_falha": "N/A",
                "motivo_transbordo": "N/A",
                "descricao": "Conversa sem conteúdo suficiente para análise",
                "sugestao_solucao": "N/A"
            }
        
        # Criar prompt
        prompt = criar_prompt_sistema(conversa)
        
        # Gerar conteúdo com retry e backoff exponencial para rate limiting
        response = None
        max_retries = 5  # Aumentado para 5 tentativas
        
        for tentativa in range(max_retries):
            try:
                response = client.chat.completions.create(
                    model=modelo,
                    messages=[
                        {"role": "system", "content": "Você é um Auditor de Qualidade de Atendimento Automatizado (QA). Retorne APENAS JSON válido, sem texto adicional."},
                        {"role": "user", "content": prompt}
                    ],
                    temperature=0.1,
                    response_format={"type": "json_object"}  # Forçar resposta JSON
                )
                break  # Sucesso, sair do loop
            except Exception as e:
                error_msg = str(e)
                error_type = type(e).__name__
                
                # Verificar se é erro de rate limit
                is_rate_limit = (
                    "429" in error_msg or 
                    "rate_limit" in error_msg.lower() or 
                    "rate limit" in error_msg.lower() or
                    "rate_limit_exceeded" in error_type or
                    "quota" in error_msg.lower() or
                    "too_many_requests" in error_msg.lower()
                )
                
                if is_rate_limit:
                    if tentativa < max_retries - 1:
                        # Backoff exponencial: 10s, 20s, 40s, 80s
                        wait_time = 10 * (2 ** tentativa)
                        # Limitar a 60 segundos máximo
                        wait_time = min(wait_time, 60)
                        
                        # Tentar extrair retry-after do header se disponível
                        if hasattr(e, 'response') and hasattr(e.response, 'headers'):
                            retry_after = e.response.headers.get('retry-after')
                            if retry_after:
                                try:
                                    wait_time = int(retry_after) + 2
                                except:
                                    pass
                        
                        time.sleep(wait_time)
                        continue  # Tentar novamente
                    else:
                        # Última tentativa falhou
                        raise Exception(f"Rate limit excedido após {max_retries} tentativas. Aguarde alguns minutos antes de tentar novamente.")
                else:
                    # Outro tipo de erro, não tentar novamente
                    raise e
        
        if response is None or not response.choices or not response.choices[0].message.content:
            return {
                "acao_necessaria": True,
                "tipo_falha": "Erro na API",
                "motivo_transbordo": "N/A",
                "descricao": "O modelo não retornou uma resposta válida",
                "sugestao_solucao": "Verificar conexão com API OpenAI e tentar novamente"
            }
        
        texto_resposta = response.choices[0].message.content.strip()
        resultado_json = extract_json_from_text(texto_resposta)
        
        if resultado_json is None:
            return {
                "acao_necessaria": True,
                "tipo_falha": "Erro ao processar resposta",
                "motivo_transbordo": "N/A",
                "descricao": f"Erro ao extrair JSON. Resposta: {texto_resposta[:150]}",
                "sugestao_solucao": "Verificar formato da resposta da API e ajustar prompt se necessário"
            }
        
        # Validar e padronizar campos
        # Processar had_need_to_transfer (novo formato) ou acao_necessaria (formato antigo para compatibilidade)
        had_need_to_transfer = resultado_json.get("had_need_to_transfer", None)
        acao_necessaria_old = resultado_json.get("acao_necessaria", None)
        
        # Converter had_need_to_transfer para acao_necessaria
        if had_need_to_transfer is not None:
            if isinstance(had_need_to_transfer, str):
                acao_necessaria = had_need_to_transfer.lower() in ["true", "sim", "yes", "1"]
            else:
                acao_necessaria = bool(had_need_to_transfer)
        elif acao_necessaria_old is not None:
            if isinstance(acao_necessaria_old, str):
                acao_necessaria = acao_necessaria_old.lower() in ["true", "sim", "yes", "1"]
            else:
                acao_necessaria = bool(acao_necessaria_old)
        else:
            acao_necessaria = False
        
        resultado_json["acao_necessaria"] = bool(acao_necessaria)
        
        # Criar tipo_falha e descricao baseados no resultado
        if acao_necessaria:
            resultado_json["tipo_falha"] = str(resultado_json.get("tipo_falha", "Necessidade de Transferência")).strip()
            resultado_json["descricao"] = str(resultado_json.get("descricao", "Conversa precisa de atenção - houve necessidade real de transferência para atendimento humano")).strip()
        else:
            resultado_json["tipo_falha"] = str(resultado_json.get("tipo_falha", "N/A")).strip()
            resultado_json["descricao"] = str(resultado_json.get("descricao", "Conversa processada corretamente - não houve necessidade de transferência")).strip()
        
        # Motivo do transbordo (sempre preencher)
        resultado_json["motivo_transbordo"] = str(resultado_json.get("motivo_transbordo", "N/A")).strip() or "N/A"
        
        # Processar sugestão de solução
        sugestao = resultado_json.get("sugestao_solucao", "")
        if not sugestao or sugestao.strip() == "":
            # Se não foi fornecida e há ação necessária, criar uma sugestão genérica
            if acao_necessaria:
                sugestao = "Revisar fluxo conversacional e melhorar detecção de casos que requerem transferência para atendimento humano"
            else:
                sugestao = "N/A"
        
        resultado_json["sugestao_solucao"] = str(sugestao).strip()
        
        return resultado_json
        
    except Exception as e:
        error_msg = str(e)
        
        # Verificar se é erro de rate limit
        is_rate_limit = (
            "429" in error_msg or 
            "quota" in error_msg.lower() or 
            "rate limit" in error_msg.lower() or 
            "rate_limit" in error_msg.lower() or
            "rate_limit_exceeded" in error_msg.lower() or
            "too_many_requests" in error_msg.lower()
        )
        
        if is_rate_limit:
            return {
                "acao_necessaria": True,
                "tipo_falha": "Rate limit excedido",
                "motivo_transbordo": "N/A",
                "descricao": "⚠️ Rate limit da API OpenAI excedido. Soluções: 1) Aumente o delay entre requisições na sidebar (recomendado: 10-15s), 2) Adicione créditos na sua conta OpenAI, 3) Aguarde alguns minutos e tente novamente.",
                "sugestao_solucao": "Aumentar delay entre requisições na sidebar para 10-15 segundos ou adicionar créditos na conta OpenAI"
            }
        
        if len(error_msg) > 200:
            error_msg = error_msg[:200] + "..."
        
        return {
            "acao_necessaria": True,
            "tipo_falha": "Erro na análise",
            "motivo_transbordo": "N/A",
            "descricao": f"Erro na análise: {error_msg}",
            "sugestao_solucao": "Verificar logs de erro e configurações da API OpenAI"
        }

# Função para analisar uma conversa localmente usando regras de negócio
def analisar_conversa_local(conversa: str) -> Dict:
    """Analisa uma conversa usando regras de negócio locais (sem API)"""
    try:
        # Verificar se a conversa não está vazia
        if not conversa or len(conversa.strip()) < 10:
            return {
                "necessidade_transbordo": "Não",
                "transferencia": "Não",
                "agente_agiu_corretamente": "Sim",
                "motivo_transbordo": "N/A",
                "problema_mapeado": "Conversa muito curta",
                "precisa_atencao": "Não",
                "observacao": "Conversa sem conteúdo suficiente para análise"
            }
        
        conversa_lower = conversa.lower()
        conversa_original = conversa
        
        # Normalizar a conversa para análise
        linhas = conversa.split('\n')
        
        # 1. NECESSIDADE DE TRANSBORDO
        necessidade_transbordo = "Não"
        motivo_transbordo = "N/A"
        
        # Padrões que indicam necessidade de transbordo
        pede_humano_patterns = [
            r'falar\s+com\s+(?:um\s+)?(?:atendente|humano|pessoa|operador)',
            r'quero\s+(?:falar\s+)?com\s+(?:um\s+)?(?:atendente|humano|pessoa)',
            r'preciso\s+de\s+(?:um\s+)?(?:atendente|humano|pessoa)',
            r'atendente\s+(?:humano|pessoa)',
            r'transferir\s+para\s+(?:um\s+)?(?:atendente|humano|pessoa)'
        ]
        
        looping_patterns = [
            r'(?:repete|repetiu|repetindo|loop)',
            r'mesma\s+(?:coisa|mensagem|resposta)',
            r'já\s+(?:falei|disse|respondi)',
            r'não\s+entende'
        ]
        
        erro_bot_patterns = [
            r'erro',
            r'não\s+funcionou',
            r'não\s+está\s+funcionando',
            r'bug',
            r'problema\s+técnico',
            r'falha'
        ]
        
        cliente_frustrado_patterns = [
            r'irritado|irritada',
            r'estou\s+bravo|estou\s+brava',
            r'não\s+resolveu',
            r'incompetente',
            r'horrível|péssimo'
        ]
        
        # Verificar necessidade de transbordo
        pede_humano = any(re.search(pattern, conversa_lower) for pattern in pede_humano_patterns)
        tem_looping = any(re.search(pattern, conversa_lower) for pattern in looping_patterns)
        tem_erro = any(re.search(pattern, conversa_lower) for pattern in erro_bot_patterns)
        cliente_frustrado = any(re.search(pattern, conversa_lower) for pattern in cliente_frustrado_patterns)
        
        # Verificar divergência (cliente nega recebimento ou status)
        divergencia_patterns = [
            r'não\s+recebi',
            r'não\s+foi\s+entregue',
            r'está\s+errado',
            r'não\s+é\s+isso',
            r'diferente\s+do\s+que\s+comprei',
            r'pedido\s+errado'
        ]
        tem_divergencia = any(re.search(pattern, conversa_lower) for pattern in divergencia_patterns)
        
        if pede_humano:
            necessidade_transbordo = "Sim"
            motivo_transbordo = "Solicitação do cliente"
        elif tem_looping:
            necessidade_transbordo = "Sim"
            motivo_transbordo = "Looping eterno"
        elif tem_divergencia:
            necessidade_transbordo = "Sim"
            motivo_transbordo = "Divergência de status"
        elif tem_erro:
            necessidade_transbordo = "Sim"
            motivo_transbordo = "Erro técnico"
        elif cliente_frustrado:
            necessidade_transbordo = "Sim"
            motivo_transbordo = "Cliente frustrado"
        
        # 2. TRANSFERÊNCIA
        transferencia = "Não"
        # Verificar se bot transferiu para fila humana (não link externo)
        transferencia_patterns = [
            r'transferindo\s+para\s+(?:um\s+)?(?:atendente|humano|equipe)',
            r'vou\s+transferir\s+você',
            r'conectando\s+com\s+(?:um\s+)?atendente'
        ]
        
        link_externo_patterns = [
            r'https?://',
            r'www\.',
            r'\.com\.br',
            r'formulário|formulario',
            r'sac|contato',
            r'troque\.app',
            r'crocs\.com\.br/contato'
        ]
        
        tem_transferencia = any(re.search(pattern, conversa_lower) for pattern in transferencia_patterns)
        tem_link = any(re.search(pattern, conversa_lower) for pattern in link_externo_patterns)
        
        if tem_transferencia and not tem_link:
            transferencia = "Sim"
        elif tem_link:
            transferencia = "Não"  # Link externo não conta como transferência
        
        # 3. AGENTE AGIU CORRETAMENTE
        agente_correto = "Sim"
        
        # Verificar problemas que indicam que o bot agiu incorretamente
        if tem_looping:
            agente_correto = "Não"
        elif tem_erro:
            agente_correto = "Não"
        elif tem_divergencia:
            agente_correto = "Não"
        
        # Verificar se bot pediu avaliação quando cliente digitou texto
        avaliacao_pattern = r'(?:avaliar|nota|avalie|de\s+1\s+a\s+5)'
        cliente_texto_antes = False
        bot_pediu_avaliacao = False
        
        for i, linha in enumerate(linhas):
            if re.search(avaliacao_pattern, linha.lower()) and ('bot' in linha.lower() or 'atendente' in linha.lower() or 'whizz' in linha.lower()):
                bot_pediu_avaliacao = True
                # Verificar se cliente digitou texto antes
                for j in range(max(0, i-3), i):
                    if 'cliente' in linhas[j].lower() and len(linhas[j]) > 20:
                        cliente_texto_antes = True
                        break
                break
        
        if bot_pediu_avaliacao and cliente_texto_antes:
            agente_correto = "Não"
        
        # 4. PROBLEMA MAPEADO
        problema_mapeado = "Tudo certo"
        
        # Padrões de problemas
        if re.search(r'pedido\s+atrasado|atrasado|demora', conversa_lower):
            problema_mapeado = "Pedido atrasado"
        elif re.search(r'entregue\s+para\s+outro|endereço\s+errado|destinatário', conversa_lower):
            problema_mapeado = "Pedido entregue para outro"
        elif re.search(r'troca|vale\s+troca|devolução', conversa_lower):
            problema_mapeado = "Dúvida Vale Troca"
        elif re.search(r'ferramenta|tool|integração', conversa_lower):
            problema_mapeado = "Falha em acionar tools"
        elif tem_looping:
            problema_mapeado = "Looping do bot"
        elif tem_erro:
            problema_mapeado = "Erro técnico"
        elif tem_divergencia:
            problema_mapeado = "Divergência de informações"
        
        # 5. PRECISA ATENÇÃO
        precisa_atencao = "Não"
        if tem_looping or tem_erro or (agente_correto == "Não" and necessidade_transbordo == "Sim"):
            precisa_atencao = "Sim"
        
        # 6. OBSERVAÇÃO - Descrição detalhada e contextualizada dos problemas encontrados
        detalhes_problemas = []
        
        # Capturar contexto específico da conversa
        contexto_cliente = []
        contexto_bot = []
        
        for linha in linhas[:20]:  # Analisar primeiras 20 linhas para contexto
            linha_lower = linha.lower()
            if 'cliente' in linha_lower and len(linha.strip()) > 15:
                contexto_cliente.append(linha.strip()[:100])
            elif any(termo in linha_lower for termo in ['bot', 'atendente', 'whizz']) and len(linha.strip()) > 15:
                contexto_bot.append(linha.strip()[:100])
        
        # Detalhar problemas específicos encontrados com contexto
        
        # Necessidade de transbordo
        if necessidade_transbordo == "Sim":
            detalhes_transbordo = [f"TRANSBORDO NECESSÁRIO - Motivo: {motivo_transbordo}"]
            
            if pede_humano:
                detalhes_transbordo.append("Cliente solicitou explicitamente atendimento humano")
            
            if tem_looping:
                # Contar respostas do bot para detectar repetição
                respostas_bot = [linha for linha in linhas if any(termo in linha.lower() for termo in ['bot', 'atendente', 'whizz'])]
                if len(respostas_bot) > 3:
                    # Verificar similaridade entre respostas
                    similar_count = 0
                    for i in range(len(respostas_bot)-1):
                        if i < len(respostas_bot)-1:
                            palavras_linha1 = set(respostas_bot[i].lower().split())
                            palavras_linha2 = set(respostas_bot[i+1].lower().split())
                            palavras_comuns = palavras_linha1 & palavras_linha2
                            if len(palavras_comuns) > 5 and len(respostas_bot[i].split()) > 5:
                                similar_count += 1
                    
                    if similar_count > 0:
                        detalhes_transbordo.append(f"Bot entrou em looping: detectadas {similar_count + 1} respostas repetitivas/conflitantes. Cliente relatou que o bot 'não entende' ou repete a mesma informação.")
                    else:
                        detalhes_transbordo.append("Bot entrou em looping - respostas repetitivas detectadas na conversa. Cliente indicou que bot repete mesma informação ou não avança no atendimento.")
                else:
                    detalhes_transbordo.append("Bot entrou em looping - padrões de repetição detectados. Cliente mencionou que bot não está entendendo ou repete respostas.")
            
            if tem_divergencia:
                if re.search(r'não\s+recebi|não\s+foi\s+entregue', conversa_lower):
                    detalhes_transbordo.append("Cliente relatou que NÃO RECEBEU o pedido, mas sistema/bot indicou como entregue - DIVERGÊNCIA CRÍTICA detectada")
                elif re.search(r'pedido\s+errado|produto\s+errado|diferente\s+do\s+que\s+comprei', conversa_lower):
                    detalhes_transbordo.append("Cliente recebeu PRODUTO/PEDIDO DIFERENTE do que solicitou - divergência entre pedido e entrega")
                elif re.search(r'está\s+errado|não\s+é\s+isso|informação\s+errada', conversa_lower):
                    detalhes_transbordo.append("Cliente contestou informações do bot dizendo que estão ERRADAS - divergência de dados/fatos")
            
            if tem_erro:
                if re.search(r'link\s+não\s+funciona|site\s+não\s+abre|não\s+consegui\s+acessar|link\s+não\s+funciona', conversa_lower):
                    detalhes_transbordo.append("ERRO TÉCNICO: Cliente relatou que link/formulário indicado pelo bot NÃO FUNCIONA. Bot direcionou para recurso inacessível.")
                elif re.search(r'erro\s+técnico|bug|falha\s+do\s+sistema|sistema\s+não\s+funciona', conversa_lower):
                    detalhes_transbordo.append("ERRO TÉCNICO detectado no sistema/bot durante a conversa")
                else:
                    detalhes_transbordo.append("Falha técnica ou erro na operação do bot detectado")
            
            if cliente_frustrado:
                detalhes_transbordo.append("Cliente demonstrou FRUSTRAÇÃO/INSATISFAÇÃO evidente durante a interação")
            
            detalhes_problemas.append(" | ".join(detalhes_transbordo))
        
        # Detalhar sobre transferência
        if transferencia == "Sim":
            detalhes_problemas.append("Bot realizou TRANSFERÊNCIA para fila humana (ação correta)")
        elif necessidade_transbordo == "Sim" and transferencia == "Não":
            if tem_link:
                detalhes_problemas.append("⚠️ PROBLEMA: Cliente precisava de transbordo, mas bot apenas direcionou para LINK EXTERNO/SAC ao invés de transferir para fila humana diretamente")
            else:
                detalhes_problemas.append("⚠️ PROBLEMA: Cliente necessitava de transbordo mas NÃO FOI TRANSFERIDO pelo bot")
        
        # Detalhar comportamento incorreto do bot
        if agente_correto == "Não":
            problemas_bot_detalhados = []
            
            if tem_looping:
                problemas_bot_detalhados.append("Bot entrou em LOOPING - repetiu mesmas respostas/mensagens, demonstrando falha no fluxo conversacional")
            
            if tem_erro:
                problemas_bot_detalhados.append("Bot apresentou ERRO TÉCNICO durante atendimento")
            
            if bot_pediu_avaliacao and cliente_texto_antes:
                problemas_bot_detalhados.append("Bot solicitou AVALIAÇÃO (nota 1-5) quando cliente havia digitado TEXTO DESCRITIVO - falha no reconhecimento de intent/fluxo")
            
            if tem_divergencia:
                problemas_bot_detalhados.append("Bot forneceu INFORMAÇÕES DIVERGENTES da realidade relatada pelo cliente")
            
            if not tem_looping and not tem_erro and not tem_divergencia and agente_correto == "Não":
                problemas_bot_detalhados.append("Bot não agiu de forma adequada para a situação do cliente")
            
            if problemas_bot_detalhados:
                detalhes_problemas.append(f"❌ BOT AGIU INCORRETAMENTE: {' | '.join(problemas_bot_detalhados)}")
        
        # Detalhar problema mapeado com contexto
        if problema_mapeado != "Tudo certo":
            detalhes_problema_mapeado = []
            
            if problema_mapeado == "Pedido atrasado":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: PEDIDO ATRASADO - Cliente está aguardando entrega que excede prazo esperado")
            elif problema_mapeado == "Pedido entregue para outro":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: PEDIDO ENTREGUE EM ENDEREÇO/DESTINATÁRIO INCORRETO - situação de logística")
            elif problema_mapeado == "Dúvida Vale Troca":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: DÚVIDA SOBRE PROCESSO DE TROCA/DEVOLUÇÃO - cliente precisa de orientação sobre política de troca")
            elif problema_mapeado == "Falha em acionar tools":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: FALHA TÉCNICA - Bot não conseguiu acionar ferramentas/integrações necessárias para resolver a demanda")
            elif problema_mapeado == "Looping do bot":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: LOOPING DO BOT - Bot ficou preso em ciclo de respostas repetitivas, não avançando no atendimento")
            elif problema_mapeado == "Erro técnico":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: ERRO TÉCNICO - Falha no sistema ou no funcionamento do bot")
            elif problema_mapeado == "Divergência de informações":
                detalhes_problema_mapeado.append("PROBLEMA MAPEADO: DIVERGÊNCIA DE INFORMAÇÕES - Dados fornecidos pelo bot não correspondem à situação real do cliente")
            
            if detalhes_problema_mapeado:
                detalhes_problemas.append(detalhes_problema_mapeado[0])
        
        # Indicar se precisa atenção especial
        if precisa_atencao == "Sim":
            detalhes_problemas.append("🚨 PRECISA ATENÇÃO ESPECIAL - Bug grave, looping ou falha crítica detectada")
        
        # Construir observação final detalhada
        if len(detalhes_problemas) > 0:
            observacao = " | ".join(detalhes_problemas)
        elif necessidade_transbordo == "Sim":
            observacao = f"Transbordo necessário: {motivo_transbordo}. Problema identificado: {problema_mapeado}. Bot {'transferiu corretamente' if transferencia == 'Sim' else 'não transferiu para fila humana'}."
        elif problema_mapeado != "Tudo certo":
            observacao = f"Problema identificado: {problema_mapeado}. Bot agiu corretamente durante o atendimento, mas há questão específica a resolver relacionada ao problema mapeado."
        else:
            observacao = "✅ Conversa processada normalmente. Bot forneceu informações adequadas, atendeu corretamente e cliente não demonstrou necessidade de transbordo ou problemas críticos."
        
        return {
            "necessidade_transbordo": necessidade_transbordo,
            "transferencia": transferencia,
            "agente_agiu_corretamente": agente_correto,
            "motivo_transbordo": motivo_transbordo,
            "problema_mapeado": problema_mapeado,
            "precisa_atencao": precisa_atencao,
            "observacao": observacao
        }
        
    except Exception as e:
        return {
            "necessidade_transbordo": "Erro",
            "transferencia": "Erro",
            "agente_agiu_corretamente": "Erro",
            "motivo_transbordo": f"Erro na análise: {str(e)[:100]}",
            "problema_mapeado": "Erro no processamento",
            "precisa_atencao": "Sim",
            "observacao": f"Erro ao analisar conversa: {str(e)[:150]}"
        }

# Função para processar arquivo TXT
def processar_txt(conteudo: str) -> List[str]:
    """Processa arquivo TXT separado por '---'"""
    conversas = []
    partes = conteudo.split("---")
    
    for parte in partes:
        parte_limpa = parte.strip()
        if parte_limpa:
            conversas.append(parte_limpa)
    
    return conversas

# Função para processar arquivo CSV
def processar_csv(conteudo: str) -> Dict:
    """Processa arquivo CSV com coluna 'conversa' ou 'Conversa' e retorna conversas + DataFrame completo"""
    try:
        # Tentar diferentes métodos de leitura do CSV
        df = None
        
        # Método 1: Tentar com engine python (melhor para células multilinha)
        try:
            df = pd.read_csv(StringIO(conteudo), quotechar='"', skipinitialspace=True, 
                           on_bad_lines='skip', engine='python', keep_default_na=False)
        except Exception as e1:
            # Método 2: Tentar com engine padrão C
            try:
                df = pd.read_csv(StringIO(conteudo), quotechar='"', skipinitialspace=True, 
                               on_bad_lines='skip', keep_default_na=False)
            except Exception as e2:
                # Método 3: Tentar sem especificar quotechar
                try:
                    df = pd.read_csv(StringIO(conteudo), skipinitialspace=True, 
                                   on_bad_lines='skip', keep_default_na=False)
                except Exception as e3:
                    # Método 4: Usar csv module manualmente (silenciosamente)
                    try:
                        import csv
                        from io import StringIO
                        reader = csv.DictReader(StringIO(conteudo))
                        rows = list(reader)
                        if rows:
                            df = pd.DataFrame(rows)
                    except Exception as e4:
                        # Se todos os métodos falharam, mostrar erro
                        pass
        
        if df is None or df.empty:
            st.error("❌ Não foi possível processar o arquivo CSV ou está vazio!")
            return {"conversas": [], "dataframe": None}
        
        # Procurar coluna de conversa (case-insensitive)
        coluna_conversa = None
        for col in df.columns:
            col_limpa = col.strip().lower()
            if col_limpa == "conversa":
                coluna_conversa = col
                break
        
        if coluna_conversa is None:
            st.error(f"❌ Coluna 'conversa' não encontrada no CSV!")
            st.info(f"📋 Colunas disponíveis no arquivo: {', '.join(df.columns.tolist()[:10])}")
            if len(df.columns) > 10:
                st.info(f"... e mais {len(df.columns) - 10} coluna(s)")
            return {"conversas": [], "dataframe": None}
        
        # Extrair conversas, removendo valores nulos e vazios
        conversas = df[coluna_conversa].dropna().tolist()
        # Converter para string e remover conversas vazias
        conversas_processadas = []
        indices_validos = []
        for idx, conv in enumerate(conversas):
            conv_str = str(conv).strip()
            if conv_str and conv_str.lower() not in ['nan', 'none', '']:
                conversas_processadas.append(conv_str)
                indices_validos.append(idx)
        
        # Filtrar DataFrame para manter apenas linhas com conversas válidas
        df_filtrado = df.iloc[indices_validos].copy() if indices_validos else df.copy()
        
        return {
            "conversas": conversas_processadas,
            "dataframe": df_filtrado
        }
    
    except Exception as e:
        st.error(f"❌ Erro ao processar CSV: {str(e)}")
        import traceback
        with st.expander("🔍 Detalhes do erro (clique para expandir)"):
            st.code(traceback.format_exc())
        return {"conversas": [], "dataframe": None}

# Interface principal
st.header("📤 Upload de Arquivo")

uploaded_file = st.file_uploader(
    "Selecione um arquivo (.txt ou .csv)",
    type=["txt", "csv"],
    help="Para .txt: conversas separadas por '---'. Para .csv: deve ter coluna 'Conversa' ou 'conversa' (case-insensitive)"
)

conversas_carregadas = []

if uploaded_file is not None:
    # Ler conteúdo do arquivo
    if uploaded_file.name.endswith('.txt'):
        try:
            conteudo = str(uploaded_file.read(), "utf-8")
            conversas_carregadas = processar_txt(conteudo)
            st.session_state['conversas_carregadas_count'] = len(conversas_carregadas)
            st.success(f"✅ {len(conversas_carregadas)} conversa(s) carregada(s) do arquivo TXT")
        except Exception as e:
            st.error(f"❌ Erro ao ler arquivo TXT: {str(e)}")
            conversas_carregadas = []
    
    elif uploaded_file.name.endswith('.csv'):
        try:
            # Tentar diferentes encodings
            bytes_data = uploaded_file.read()
            conteudo = None
            
            for encoding in ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252']:
                try:
                    conteudo = bytes_data.decode(encoding)
                    break
                except:
                    continue
            
            if conteudo is None:
                conteudo = bytes_data.decode('utf-8', errors='ignore')
            
            resultado_csv = processar_csv(conteudo)
            conversas_carregadas = resultado_csv.get("conversas", [])
            df_original = resultado_csv.get("dataframe", None)
            
            # Salvar DataFrame original no session state
            if df_original is not None:
                st.session_state['df_csv_original'] = df_original
            else:
                st.session_state['df_csv_original'] = None
            
            if conversas_carregadas:
                st.session_state['conversas_carregadas_count'] = len(conversas_carregadas)
                st.success(f"✅ {len(conversas_carregadas)} conversa(s) carregada(s) do arquivo CSV")
                # Mostrar prévia da primeira conversa para debug
                if len(conversas_carregadas) > 0:
                    st.info(f"📝 Prévia da primeira conversa (primeiros 300 caracteres): {conversas_carregadas[0][:300]}...")
            else:
                st.warning("⚠️ Nenhuma conversa foi encontrada no arquivo CSV. Verifique se a coluna 'Conversa' existe.")
        except Exception as e:
            st.error(f"❌ Erro ao ler arquivo CSV: {str(e)}")
            import traceback
            st.error(f"Detalhes: {traceback.format_exc()}")
            conversas_carregadas = []
            st.session_state['df_csv_original'] = None
    
    # Mostrar prévia das conversas e informações sobre limite
    if conversas_carregadas:
        # Mostrar informações sobre quantas conversas serão analisadas
        limite = st.session_state.get('limite_conversas', None)
        total_carregadas = len(conversas_carregadas)
        
        if limite and limite < total_carregadas:
            st.warning(f"⚠️ **Limite configurado**: Das {total_carregadas} conversas carregadas, apenas as primeiras **{limite}** serão analisadas. Para analisar todas, deixe o campo 'Número máximo de conversas' vazio na sidebar.")
        else:
            st.success(f"✅ **Todas as {total_carregadas} conversas** serão analisadas.")
        
        with st.expander("👁️ Prévia das conversas carregadas"):
            for idx, conversa in enumerate(conversas_carregadas[:3], 1):
                st.markdown(f"**Conversa {idx}:**")
                st.text(conversa[:500] + "..." if len(conversa) > 500 else conversa)
                st.markdown("---")
            
            if len(conversas_carregadas) > 3:
                st.info(f"*E mais {len(conversas_carregadas) - 3} conversa(s)...*")

# Função wrapper para análise via OpenAI
def analisar_conversa(conversa: str, modelo: str, api_key_openai: str) -> Dict:
    """Analisa uma conversa usando OpenAI API"""
    if modelo is None:
        return {
            "acao_necessaria": True,
            "tipo_falha": "Erro de configuração",
            "motivo_transbordo": "N/A",
            "descricao": "Erro: Modelo OpenAI não foi especificado",
            "sugestao_solucao": "Selecionar um modelo OpenAI na barra lateral"
        }
    return analisar_conversa_openai(conversa, modelo, api_key_openai)

# Processamento
st.header("🔄 Processamento")

# Mostrar método selecionado
if not api_key:
    st.warning("⚠️ Por favor, configure a OpenAI API Key na barra lateral antes de iniciar a análise.")
else:
    st.info(f"🔍 **Análise via IA (OpenAI)** | **Modelo:** {model_name}")
    
    # Aviso sobre rate limits se houver muitas conversas
    if conversas_carregadas and len(conversas_carregadas) > 50:
        st.warning(f"⚠️ **Atenção**: Você tem {len(conversas_carregadas)} conversas para analisar. Para evitar rate limits, recomendamos:")
        st.markdown("""
        - **Aumentar o delay entre requisições** na sidebar (10-15 segundos para contas gratuitas)
        - **Verificar créditos** na sua conta OpenAI (platform.openai.com)
        - **Aguarde alguns minutos** se receber erros de rate limit
        """)

if conversas_carregadas and st.button("🚀 Iniciar Análise", type="primary", use_container_width=True):
    if len(conversas_carregadas) == 0:
        st.error("❌ Nenhuma conversa encontrada para analisar!")
    else:
        # Verificar se API Key está configurada
        if not api_key:
            st.error("❌ Por favor, configure a OpenAI API Key na barra lateral!")
            st.stop()
        
        # Aplicar limite de conversas se configurado
        limite = st.session_state.get('limite_conversas', None)
        conversas_para_analisar = conversas_carregadas
        
        if limite and limite < len(conversas_carregadas):
            conversas_para_analisar = conversas_carregadas[:limite]
            st.info(f"📊 **Limite aplicado**: Analisando apenas as primeiras {limite} de {len(conversas_carregadas)} conversas carregadas.")
        else:
            st.info(f"📊 Analisando todas as {len(conversas_carregadas)} conversas carregadas.")
        
        # Salvar conversas para análise no session state para garantir acesso posterior
        st.session_state['conversas_para_analisar'] = conversas_para_analisar
        
        # Inicializar lista de resultados
        resultados = []
        
        # Barra de progresso
        progress_bar = st.progress(0)
        status_text = st.empty()
        
        # Obter DataFrame original se disponível
        df_original = st.session_state.get('df_csv_original', None)
        
        # Função auxiliar para encontrar coluna por nome (case-insensitive)
        def encontrar_coluna(df, nomes_possiveis):
            if df is None:
                return None
            for nome in nomes_possiveis:
                for col in df.columns:
                    if col.strip().lower() == nome.lower():
                        return col
            return None
        
        # Identificar colunas relevantes no CSV original
        col_retailer = encontrar_coluna(df_original, ['retailer', 'cliente', 'customer', 'loja', 'store'])
        col_data = encontrar_coluna(df_original, ['data', 'date', 'data_hora', 'datetime', 'timestamp'])
        col_hora = encontrar_coluna(df_original, ['hora', 'time', 'horario'])
        col_conversa_original = encontrar_coluna(df_original, ['conversa', 'conversation'])
        col_csr_id = encontrar_coluna(df_original, ['csr id', 'csr_id', 'csrid', 'csr', 'atendente id', 'atendente_id'])
        col_chat_id = encontrar_coluna(df_original, ['chat id', 'chat_id', 'chatid', 'chat', 'conversation id', 'conversation_id'])
        
        # Iterar sobre as conversas (limitadas)
        total_conversas = len(conversas_para_analisar)
        for idx, conversa in enumerate(conversas_para_analisar, 1):
            status_text.text(f"📊 Analisando conversa {idx}/{total_conversas} (OpenAI API)...")
            
            # Analisar conversa usando OpenAI API
            resultado = analisar_conversa(conversa, model_name, api_key)
            # Delay configurável para evitar rate limiting
            time.sleep(delay_entre_requisicoes)
            
            resultado["conversa_numero"] = idx
            resultado["conversa"] = conversa[:200] + "..." if len(conversa) > 200 else conversa
            resultado["conversa_completa"] = conversa  # Manter conversa completa para download
            
            # Adicionar informações do CSV original se disponível
            if df_original is not None and idx <= len(df_original):
                linha_original = df_original.iloc[idx - 1]
                
                # Adicionar retailer/cliente
                if col_retailer:
                    resultado["retailer"] = str(linha_original.get(col_retailer, "N/A")).strip()
                else:
                    resultado["retailer"] = "N/A"
                
                # Adicionar data
                if col_data:
                    data_valor = linha_original.get(col_data, "N/A")
                    resultado["data"] = str(data_valor).strip() if pd.notna(data_valor) else "N/A"
                else:
                    resultado["data"] = "N/A"
                
                # Adicionar hora (se coluna separada)
                if col_hora:
                    hora_valor = linha_original.get(col_hora, "N/A")
                    resultado["hora"] = str(hora_valor).strip() if pd.notna(hora_valor) else "N/A"
                else:
                    resultado["hora"] = "N/A"
                
                # Adicionar csr id
                if col_csr_id:
                    csr_id_valor = linha_original.get(col_csr_id, "N/A")
                    resultado["csr_id"] = str(csr_id_valor).strip() if pd.notna(csr_id_valor) else "N/A"
                else:
                    resultado["csr_id"] = "N/A"
                
                # Adicionar chat id
                if col_chat_id:
                    chat_id_valor = linha_original.get(col_chat_id, "N/A")
                    resultado["chat_id"] = str(chat_id_valor).strip() if pd.notna(chat_id_valor) else "N/A"
                else:
                    resultado["chat_id"] = "N/A"
            else:
                resultado["retailer"] = "N/A"
                resultado["data"] = "N/A"
                resultado["hora"] = "N/A"
                resultado["csr_id"] = "N/A"
                resultado["chat_id"] = "N/A"
            
            resultados.append(resultado)
            
            # Atualizar progresso
            progress = idx / total_conversas
            progress_bar.progress(progress)
        
        status_text.text("✅ Análise concluída!")
        
        # Criar DataFrame com resultados
        df_resultados = pd.DataFrame(resultados)
        
        # Garantir que colunas essenciais existam (adicionar se não estiverem presentes)
        if "sugestao_solucao" not in df_resultados.columns:
            df_resultados["sugestao_solucao"] = "N/A"
        if "retailer" not in df_resultados.columns:
            df_resultados["retailer"] = "N/A"
        if "data" not in df_resultados.columns:
            df_resultados["data"] = "N/A"
        if "hora" not in df_resultados.columns:
            df_resultados["hora"] = "N/A"
        if "csr_id" not in df_resultados.columns:
            df_resultados["csr_id"] = "N/A"
        if "chat_id" not in df_resultados.columns:
            df_resultados["chat_id"] = "N/A"
        if "conversa_completa" not in df_resultados.columns:
            # Se não existe, tentar recriar a partir das conversas originais
            # Isso pode acontecer se houver algum problema no processamento
            df_resultados["conversa_completa"] = ""
        if "motivo_transbordo" not in df_resultados.columns:
            df_resultados["motivo_transbordo"] = "N/A"
        
        # Garantir que conversa_completa não seja uma versão resumida
        # Se conversa_completa está vazia ou é igual à conversa resumida, tentar recuperar da lista original
        conversas_originais = st.session_state.get('conversas_para_analisar', conversas_para_analisar if 'conversas_para_analisar' in locals() else [])
        
        if "conversa_completa" in df_resultados.columns and "conversa" in df_resultados.columns:
            for idx in df_resultados.index:
                conv_completa = str(df_resultados.loc[idx, "conversa_completa"]).strip()
                conv_resumida = str(df_resultados.loc[idx, "conversa"]).strip()
                
                # Se conversa_completa está vazia ou é igual à resumida (e a resumida termina com "...")
                if not conv_completa or (conv_resumida.endswith("...") and conv_completa == conv_resumida):
                    # Tentar recuperar da lista de conversas originais
                    num_conversa = df_resultados.loc[idx, "conversa_numero"] if "conversa_numero" in df_resultados.columns else None
                    if num_conversa and isinstance(num_conversa, (int, float)) and conversas_originais and int(num_conversa) <= len(conversas_originais):
                        idx_original = int(num_conversa) - 1
                        if 0 <= idx_original < len(conversas_originais):
                            df_resultados.loc[idx, "conversa_completa"] = conversas_originais[idx_original]
                    elif conv_resumida and not conv_resumida.endswith("..."):
                        # Se a conversa não está resumida, usar ela como completa
                        df_resultados.loc[idx, "conversa_completa"] = conv_resumida
        
        # Preencher valores vazios
        df_resultados["sugestao_solucao"] = df_resultados["sugestao_solucao"].fillna("N/A")
        df_resultados["retailer"] = df_resultados["retailer"].fillna("N/A")
        df_resultados["data"] = df_resultados["data"].fillna("N/A")
        df_resultados["hora"] = df_resultados["hora"].fillna("N/A")
        df_resultados["csr_id"] = df_resultados["csr_id"].fillna("N/A") if "csr_id" in df_resultados.columns else "N/A"
        df_resultados["chat_id"] = df_resultados["chat_id"].fillna("N/A") if "chat_id" in df_resultados.columns else "N/A"
        df_resultados["motivo_transbordo"] = df_resultados["motivo_transbordo"].fillna("N/A") if "motivo_transbordo" in df_resultados.columns else "N/A"
        df_resultados["conversa_completa"] = df_resultados["conversa_completa"].fillna("")
        
        # Garantir que conversa_completa nunca seja vazia - usar conversa como fallback
        if "conversa" in df_resultados.columns:
            mask_vazia = (df_resultados["conversa_completa"].isna()) | (df_resultados["conversa_completa"].astype(str).str.strip() == "")
            df_resultados.loc[mask_vazia, "conversa_completa"] = df_resultados.loc[mask_vazia, "conversa"]
        
        # Reordenar colunas
        colunas_ordenadas = [
            "conversa_numero",
            "retailer",
            "data",
            "hora",
            "csr_id",
            "chat_id",
            "acao_necessaria",
            "tipo_falha",
            "motivo_transbordo",
            "descricao",
            "sugestao_solucao",
            "conversa"
        ]
        
        # Verificar se todas as colunas existem antes de reordenar
        colunas_existentes = [col for col in colunas_ordenadas if col in df_resultados.columns]
        if len(colunas_existentes) == len(colunas_ordenadas):
            df_resultados = df_resultados[colunas_ordenadas]
        
        # Salvar no session state
        st.session_state['df_resultados'] = df_resultados
        st.session_state['resultados_processados'] = True

# Exibição dos resultados
if 'resultados_processados' in st.session_state and st.session_state['resultados_processados']:
    st.header("📊 Resultados da Análise")
    
    df_resultados = st.session_state['df_resultados']
    
    # Estatísticas rápidas
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.metric("Total de Conversas", len(df_resultados))
    
    with col2:
        # Converter acao_necessaria para boolean se necessário
        if "acao_necessaria" in df_resultados.columns:
            acoes_necessarias = df_resultados["acao_necessaria"].apply(
                lambda x: x if isinstance(x, bool) else str(x).lower() in ["true", "sim", "yes", "1"]
            ).sum()
        else:
            acoes_necessarias = 0
        st.metric("Ações Necessárias", acoes_necessarias, delta=None)
    
    with col3:
        if "acao_necessaria" in df_resultados.columns:
            acoes_necessarias = df_resultados["acao_necessaria"].apply(
                lambda x: x if isinstance(x, bool) else str(x).lower() in ["true", "sim", "yes", "1"]
            ).sum()
            sem_acao = len(df_resultados) - acoes_necessarias
            st.metric("Sem Ação Necessária", sem_acao, delta=f"{sem_acao/len(df_resultados)*100:.1f}%")
        else:
            st.metric("Sem Ação Necessária", 0)
    
    # Dataframe com destaque
    st.subheader("Tabela de Resultados")
    
    # Preparar dataframe para exibição com destaque
    df_display = df_resultados.copy()
    
    # Exibir dataframe
    st.dataframe(
        df_display,
        use_container_width=True,
        height=400,
        hide_index=True
    )
    
    # Seção de Sugestões de Solução
    if "acao_necessaria" in df_display.columns and "sugestao_solucao" in df_display.columns:
        # Filtrar conversas que precisam de ação e têm sugestão
        df_com_sugestoes = df_display[
            df_display['acao_necessaria'].apply(
                lambda x: x if isinstance(x, bool) else str(x).lower() in ["true", "sim", "yes", "1"]
            ) & 
            (df_display['sugestao_solucao'] != "N/A") &
            (df_display['sugestao_solucao'].notna()) &
            (df_display['sugestao_solucao'].str.strip() != "")
        ]
        
        if not df_com_sugestoes.empty:
            st.subheader("💡 Sugestões de Solução")
            st.info(f"📋 Encontradas {len(df_com_sugestoes)} conversa(s) com problemas e sugestões de solução:")
            
            for idx, row in df_com_sugestoes.iterrows():
                with st.expander(f"🔧 Conversa #{row.get('conversa_numero', idx)} - {row.get('tipo_falha', 'Problema identificado')}"):
                    st.markdown(f"**Problema:** {row.get('descricao', 'N/A')}")
                    st.markdown(f"**💡 Sugestão de Solução:** {row.get('sugestao_solucao', 'N/A')}")
    
    # Filtro para destacar conversas que precisam ação
    st.info("💡 **Dica:** Use o filtro abaixo para visualizar apenas as conversas que requerem ação/intervenção.")
    
    filtro_acao = st.checkbox("Mostrar apenas conversas que requerem ação", value=False)
    
    if filtro_acao:
        if "acao_necessaria" in df_display.columns:
            df_filtrado = df_display[df_display['acao_necessaria'].apply(
                lambda x: x if isinstance(x, bool) else str(x).lower() in ["true", "sim", "yes", "1"]
            )]
            if not df_filtrado.empty:
                st.dataframe(
                    df_filtrado,
                    use_container_width=True,
                    height=300,
                    hide_index=True
                )
                st.warning(f"⚠️ {len(df_filtrado)} conversa(s) requer(em) ação/intervenção!")
            else:
                st.success("✅ Nenhuma conversa requer ação especial!")
        else:
            st.warning("⚠️ Campo 'acao_necessaria' não encontrado nos resultados.")
    
    # Botões de download
    st.subheader("💾 Download do Relatório")
    
    # Validação: Comparar conversas originais com conversas no DataFrame final
    st.markdown("### ✅ Validação de Integridade das Conversas")
    
    conversas_originais = st.session_state.get('conversas_para_analisar', [])
    problemas_encontrados = []
    conversas_validadas = 0
    
    if conversas_originais and "conversa_completa" in df_resultados.columns and "conversa_numero" in df_resultados.columns:
        for idx in df_resultados.index:
            num_conversa = df_resultados.loc[idx, "conversa_numero"]
            if isinstance(num_conversa, (int, float)) and int(num_conversa) <= len(conversas_originais):
                idx_original = int(num_conversa) - 1
                if 0 <= idx_original < len(conversas_originais):
                    conversa_original = str(conversas_originais[idx_original])
                    conversa_no_df = str(df_resultados.loc[idx, "conversa_completa"])
                    
                    # Comparar número de caracteres
                    chars_original = len(conversa_original)
                    chars_no_df = len(conversa_no_df)
                    
                    if chars_original != chars_no_df:
                        problemas_encontrados.append({
                            "conversa_numero": int(num_conversa),
                            "chars_original": chars_original,
                            "chars_no_df": chars_no_df,
                            "diferenca": chars_original - chars_no_df
                        })
                    else:
                        conversas_validadas += 1
        
        # Exibir resultado da validação
        if problemas_encontrados:
            st.error(f"❌ **ATENÇÃO**: {len(problemas_encontrados)} conversa(s) com diferença no número de caracteres detectada(s)!")
            with st.expander("🔍 Detalhes das conversas com problema", expanded=False):
                df_problemas = pd.DataFrame(problemas_encontrados)
                st.dataframe(df_problemas, use_container_width=True, hide_index=True)
                st.warning("⚠️ As conversas serão corrigidas automaticamente antes do download.")
            
            # Corrigir automaticamente as conversas com problema
            for problema in problemas_encontrados:
                num_conv = problema["conversa_numero"]
                idx_original = num_conv - 1
                if 0 <= idx_original < len(conversas_originais):
                    # Encontrar a linha no DataFrame
                    mask = df_resultados["conversa_numero"] == num_conv
                    if mask.any():
                        df_resultados.loc[mask, "conversa_completa"] = conversas_originais[idx_original]
                        st.info(f"✅ Conversa #{num_conv} corrigida: {problema['chars_original']} caracteres restaurados.")
        else:
            st.success(f"✅ **Validação concluída**: Todas as {conversas_validadas} conversa(s) têm o mesmo número de caracteres da conversa original analisada!")
    else:
        st.info("ℹ️ Validação não disponível: conversas originais não encontradas no session state.")
    
    st.markdown("---")
    
    # Seção de download filtrado - Conversas que precisam atenção
    st.markdown("### 🔴 Download Filtrado - Conversas que Precisam Atenção")
    st.info("💡 Baixe apenas as conversas que requerem ação/intervenção, agrupadas por cliente (retailer), com conversas completas.")
    
    # Filtrar conversas que precisam atenção
    if "acao_necessaria" in df_resultados.columns:
        df_com_atencao = df_resultados[
            df_resultados['acao_necessaria'].apply(
                lambda x: x if isinstance(x, bool) else str(x).lower() in ["true", "sim", "yes", "1"]
            )
        ].copy()
        
        if not df_com_atencao.empty:
            # Preparar DataFrame para download filtrado
            df_download_filtrado = df_com_atencao.copy()
            
            # Garantir que conversa_completa contém a conversa completa original
            # Se conversa_completa não existe ou está truncada, recuperar da lista original
            conversas_originais = st.session_state.get('conversas_para_analisar', [])
            
            if "conversa_completa" in df_download_filtrado.columns:
                # Verificar e corrigir cada linha para garantir conversa completa
                for idx in df_download_filtrado.index:
                    conv_completa_atual = str(df_download_filtrado.loc[idx, "conversa_completa"])
                    conv_resumida = str(df_download_filtrado.loc[idx, "conversa"]) if "conversa" in df_download_filtrado.columns else ""
                    
                    # Se conversa_completa está vazia, truncada ou igual à resumida, recuperar original
                    if (not conv_completa_atual or 
                        conv_completa_atual.endswith("...") or 
                        (conv_resumida.endswith("...") and conv_completa_atual == conv_resumida)):
                        # Tentar recuperar da lista original
                        num_conversa = df_download_filtrado.loc[idx, "conversa_numero"] if "conversa_numero" in df_download_filtrado.columns else None
                        if num_conversa and isinstance(num_conversa, (int, float)) and conversas_originais and int(num_conversa) <= len(conversas_originais):
                            idx_original = int(num_conversa) - 1
                            if 0 <= idx_original < len(conversas_originais):
                                df_download_filtrado.loc[idx, "conversa_completa"] = conversas_originais[idx_original]
                
                # Substituir conversa resumida pela conversa completa
                df_download_filtrado["conversa"] = df_download_filtrado["conversa_completa"].astype(str)
            
            # Validação final: Verificar se todas as conversas têm o mesmo número de caracteres da original
            conversas_originais_validacao = st.session_state.get('conversas_para_analisar', [])
            if conversas_originais_validacao and "conversa_numero" in df_download_filtrado.columns:
                for idx in df_download_filtrado.index:
                    num_conversa = df_download_filtrado.loc[idx, "conversa_numero"]
                    if isinstance(num_conversa, (int, float)) and int(num_conversa) <= len(conversas_originais_validacao):
                        idx_original = int(num_conversa) - 1
                        if 0 <= idx_original < len(conversas_originais_validacao):
                            conversa_original = str(conversas_originais_validacao[idx_original])
                            conversa_no_df = str(df_download_filtrado.loc[idx, "conversa_completa"])
                            
                            # Se número de caracteres diferente, corrigir
                            if len(conversa_original) != len(conversa_no_df):
                                df_download_filtrado.loc[idx, "conversa_completa"] = conversa_original
            
            # Selecionar colunas para download (remover conversa_completa se existir)
            colunas_download = [col for col in df_download_filtrado.columns if col != "conversa_completa"]
            df_download_filtrado = df_download_filtrado[colunas_download]
            
            # Ordenar por retailer (cliente) e depois por data/hora se disponível
            colunas_ordenacao = []
            if "retailer" in df_download_filtrado.columns:
                colunas_ordenacao.append("retailer")
            if "data" in df_download_filtrado.columns:
                colunas_ordenacao.append("data")
            if "hora" in df_download_filtrado.columns:
                colunas_ordenacao.append("hora")
            
            if colunas_ordenacao:
                df_download_filtrado = df_download_filtrado.sort_values(by=colunas_ordenacao)
            
            # Reordenar colunas para download
            colunas_ordenadas_download = [
                "retailer",
                "data",
                "hora",
                "csr_id",
                "chat_id",
                "conversa_numero",
                "acao_necessaria",
                "tipo_falha",
                "motivo_transbordo",
                "descricao",
                "sugestao_solucao",
                "conversa"
            ]
            
            # Manter apenas colunas que existem
            colunas_finais = [col for col in colunas_ordenadas_download if col in df_download_filtrado.columns]
            # Adicionar outras colunas que não estão na lista
            outras_colunas = [col for col in df_download_filtrado.columns if col not in colunas_finais]
            colunas_finais = colunas_finais + outras_colunas
            
            df_download_filtrado = df_download_filtrado[colunas_finais]
            
            st.success(f"✅ {len(df_download_filtrado)} conversa(s) que precisam de atenção encontrada(s).")
            
            # Validação final antes do download: Comparar número de caracteres
            conversas_originais_final = st.session_state.get('conversas_para_analisar', [])
            validacao_final = []
            if conversas_originais_final and "conversa_numero" in df_download_filtrado.columns and "conversa" in df_download_filtrado.columns:
                for idx in df_download_filtrado.index:
                    num_conversa = df_download_filtrado.loc[idx, "conversa_numero"]
                    if isinstance(num_conversa, (int, float)) and int(num_conversa) <= len(conversas_originais_final):
                        idx_original = int(num_conversa) - 1
                        if 0 <= idx_original < len(conversas_originais_final):
                            conversa_original = str(conversas_originais_final[idx_original])
                            conversa_no_df = str(df_download_filtrado.loc[idx, "conversa"])
                            chars_original = len(conversa_original)
                            chars_no_df = len(conversa_no_df)
                            validacao_final.append({
                                "conversa": int(num_conversa),
                                "chars_original": chars_original,
                                "chars_planilha": chars_no_df,
                                "igual": chars_original == chars_no_df
                            })
            
            # Exibir resumo da validação final
            if validacao_final:
                todas_iguais = all(v["igual"] for v in validacao_final)
                if todas_iguais:
                    st.success(f"✅ **Validação Final**: Todas as {len(validacao_final)} conversa(s) na planilha têm o mesmo número de caracteres da conversa analisada!")
                else:
                    diferentes = [v for v in validacao_final if not v["igual"]]
                    st.error(f"❌ **ATENÇÃO**: {len(diferentes)} conversa(s) com diferença detectada! Corrigindo automaticamente...")
                    # Corrigir automaticamente
                    for v in diferentes:
                        num_conv = v["conversa"]
                        idx_original = num_conv - 1
                        if 0 <= idx_original < len(conversas_originais_final):
                            mask = df_download_filtrado["conversa_numero"] == num_conv
                            if mask.any():
                                df_download_filtrado.loc[mask, "conversa"] = conversas_originais_final[idx_original]
                    st.success("✅ Conversas corrigidas automaticamente!")
            
            col_filtrado1, col_filtrado2 = st.columns(2)
            
            with col_filtrado1:
                # CSV filtrado - garantir que conversa seja string completa
                # Converter conversa para string explícita para evitar truncamento
                if "conversa" in df_download_filtrado.columns:
                    df_download_filtrado["conversa"] = df_download_filtrado["conversa"].astype(str)
                
                # Salvar CSV sem limitações
                csv_filtrado = df_download_filtrado.to_csv(
                    index=False,
                    quoting=csv.QUOTE_ALL  # QUOTE_ALL para garantir que conversas com vírgulas sejam preservadas
                ).encode('utf-8-sig')
                
                st.download_button(
                    label="📥 Download CSV (Filtrado)",
                    data=csv_filtrado,
                    file_name=f"conversas_atencao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    key="download_csv_filtrado"
                )
            
            with col_filtrado2:
                # Excel filtrado - criar em memória
                # Garantir que conversa seja string completa
                if "conversa" in df_download_filtrado.columns:
                    df_download_filtrado["conversa"] = df_download_filtrado["conversa"].astype(str)
                
                # Importar openpyxl para ajustar formatação
                try:
                    import openpyxl
                    from openpyxl.styles import Alignment
                except ImportError:
                    openpyxl = None
                    Alignment = None
                
                excel_buffer_filtrado = BytesIO()
                with pd.ExcelWriter(excel_buffer_filtrado, engine='openpyxl') as writer:
                    # Agrupar por retailer se disponível
                    if "retailer" in df_download_filtrado.columns and df_download_filtrado["retailer"].nunique() > 1:
                        retailers = df_download_filtrado["retailer"].unique()
                        for retailer in retailers:
                            if pd.notna(retailer) and str(retailer).strip() != "N/A":
                                df_retailer = df_download_filtrado[df_download_filtrado["retailer"] == retailer]
                                sheet_name = str(retailer)[:31]  # Limite de 31 caracteres para nome da aba
                                df_retailer.to_excel(writer, index=False, sheet_name=sheet_name)
                                
                                # Ajustar largura da coluna de conversa para não truncar
                                worksheet = writer.sheets[sheet_name]
                                if "conversa" in df_retailer.columns:
                                    col_idx = df_retailer.columns.get_loc("conversa") + 1
                                    worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 100
                                    # Habilitar quebra de texto
                                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                                        for cell in row:
                                            if Alignment:
                                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            else:
                                # Conversas sem retailer definido
                                df_sem_retailer = df_download_filtrado[
                                    (df_download_filtrado["retailer"].isna()) | 
                                    (df_download_filtrado["retailer"].astype(str).str.strip() == "N/A")
                                ]
                                if not df_sem_retailer.empty:
                                    df_sem_retailer.to_excel(writer, index=False, sheet_name="Sem Retailer")
                                    # Ajustar largura da coluna de conversa
                                    worksheet = writer.sheets["Sem Retailer"]
                                    if "conversa" in df_sem_retailer.columns:
                                        col_idx = df_sem_retailer.columns.get_loc("conversa") + 1
                                        worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 100
                                        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                                            for cell in row:
                                                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                    else:
                        # Se não há retailer ou apenas um, criar uma única aba
                        df_download_filtrado.to_excel(writer, index=False, sheet_name='Conversas Atenção')
                        # Ajustar largura da coluna de conversa
                        worksheet = writer.sheets['Conversas Atenção']
                        if "conversa" in df_download_filtrado.columns:
                            col_idx = df_download_filtrado.columns.get_loc("conversa") + 1
                            worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 100
                            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                                for cell in row:
                                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                
                excel_data_filtrado = excel_buffer_filtrado.getvalue()
                
                st.download_button(
                    label="📥 Download Excel (Filtrado)",
                    data=excel_data_filtrado,
                    file_name=f"conversas_atencao_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_excel_filtrado"
                )
        else:
            st.info("ℹ️ Nenhuma conversa precisa de atenção. Todos os downloads abaixo incluem todas as conversas.")
    else:
        st.warning("⚠️ Campo 'acao_necessaria' não encontrado. Downloads abaixo incluem todas as conversas.")
    
    st.markdown("---")
    st.markdown("### 📊 Download Completo - Todas as Conversas")
    
    # Preparar DataFrame para download completo com conversas completas
    df_download_completo = df_resultados.copy()
    
    # Garantir que conversa_completa contém a conversa completa original
    # Se conversa_completa não existe ou está truncada, recuperar da lista original
    conversas_originais = st.session_state.get('conversas_para_analisar', [])
    
    if "conversa_completa" in df_download_completo.columns:
        # Verificar e corrigir cada linha para garantir conversa completa
        for idx in df_download_completo.index:
            conv_completa_atual = str(df_download_completo.loc[idx, "conversa_completa"])
            conv_resumida = str(df_download_completo.loc[idx, "conversa"]) if "conversa" in df_download_completo.columns else ""
            
            # Se conversa_completa está vazia, truncada ou igual à resumida, recuperar original
            if (not conv_completa_atual or 
                conv_completa_atual.endswith("...") or 
                (conv_resumida.endswith("...") and conv_completa_atual == conv_resumida)):
                # Tentar recuperar da lista original
                num_conversa = df_download_completo.loc[idx, "conversa_numero"] if "conversa_numero" in df_download_completo.columns else None
                if num_conversa and isinstance(num_conversa, (int, float)) and conversas_originais and int(num_conversa) <= len(conversas_originais):
                    idx_original = int(num_conversa) - 1
                    if 0 <= idx_original < len(conversas_originais):
                        df_download_completo.loc[idx, "conversa_completa"] = conversas_originais[idx_original]
        
        # Validação final: Verificar se todas as conversas têm o mesmo número de caracteres da original
        conversas_originais_validacao = st.session_state.get('conversas_para_analisar', [])
        if conversas_originais_validacao and "conversa_numero" in df_download_completo.columns:
            for idx in df_download_completo.index:
                num_conversa = df_download_completo.loc[idx, "conversa_numero"]
                if isinstance(num_conversa, (int, float)) and int(num_conversa) <= len(conversas_originais_validacao):
                    idx_original = int(num_conversa) - 1
                    if 0 <= idx_original < len(conversas_originais_validacao):
                        conversa_original = str(conversas_originais_validacao[idx_original])
                        conversa_no_df = str(df_download_completo.loc[idx, "conversa_completa"])
                        
                        # Se número de caracteres diferente, corrigir
                        if len(conversa_original) != len(conversa_no_df):
                            df_download_completo.loc[idx, "conversa_completa"] = conversa_original
        
        # Substituir conversa resumida pela conversa completa
        df_download_completo["conversa"] = df_download_completo["conversa_completa"].astype(str)
    
    # Remover coluna conversa_completa se existir (já foi copiada para conversa)
    if "conversa_completa" in df_download_completo.columns:
        df_download_completo = df_download_completo.drop(columns=["conversa_completa"])
    
    # Validação final antes do download completo: Comparar número de caracteres
    conversas_originais_final_completo = st.session_state.get('conversas_para_analisar', [])
    validacao_final_completo = []
    if conversas_originais_final_completo and "conversa_numero" in df_download_completo.columns and "conversa" in df_download_completo.columns:
        for idx in df_download_completo.index:
            num_conversa = df_download_completo.loc[idx, "conversa_numero"]
            if isinstance(num_conversa, (int, float)) and int(num_conversa) <= len(conversas_originais_final_completo):
                idx_original = int(num_conversa) - 1
                if 0 <= idx_original < len(conversas_originais_final_completo):
                    conversa_original = str(conversas_originais_final_completo[idx_original])
                    conversa_no_df = str(df_download_completo.loc[idx, "conversa"])
                    chars_original = len(conversa_original)
                    chars_no_df = len(conversa_no_df)
                    validacao_final_completo.append({
                        "conversa": int(num_conversa),
                        "chars_original": chars_original,
                        "chars_planilha": chars_no_df,
                        "igual": chars_original == chars_no_df
                    })
    
    # Exibir resumo da validação final
    if validacao_final_completo:
        todas_iguais = all(v["igual"] for v in validacao_final_completo)
        if todas_iguais:
            st.success(f"✅ **Validação Final**: Todas as {len(validacao_final_completo)} conversa(s) na planilha têm o mesmo número de caracteres da conversa analisada!")
        else:
            diferentes = [v for v in validacao_final_completo if not v["igual"]]
            st.error(f"❌ **ATENÇÃO**: {len(diferentes)} conversa(s) com diferença detectada! Corrigindo automaticamente...")
            # Corrigir automaticamente
            for v in diferentes:
                num_conv = v["conversa"]
                idx_original = num_conv - 1
                if 0 <= idx_original < len(conversas_originais_final_completo):
                    mask = df_download_completo["conversa_numero"] == num_conv
                    if mask.any():
                        df_download_completo.loc[mask, "conversa"] = conversas_originais_final_completo[idx_original]
            st.success("✅ Conversas corrigidas automaticamente!")
    
    col1, col2 = st.columns(2)
    
    with col1:
        # CSV - garantir que conversa seja string completa
        if "conversa" in df_download_completo.columns:
            df_download_completo["conversa"] = df_download_completo["conversa"].astype(str)
        
        # Salvar CSV sem limitações
        csv = df_download_completo.to_csv(
            index=False,
            quoting=csv.QUOTE_ALL  # QUOTE_ALL para garantir que conversas com vírgulas sejam preservadas
        ).encode('utf-8-sig')
        
        st.download_button(
            label="📥 Download CSV (Completo)",
            data=csv,
            file_name=f"relatorio_qa_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv",
            use_container_width=True
        )
    
    with col2:
        # Excel - criar em memória
        # Garantir que conversa seja string completa
        if "conversa" in df_download_completo.columns:
            df_download_completo["conversa"] = df_download_completo["conversa"].astype(str)
        
        # Importar openpyxl para ajustar formatação
        try:
            import openpyxl
            from openpyxl.styles import Alignment
        except ImportError:
            openpyxl = None
            Alignment = None
        
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df_download_completo.to_excel(writer, index=False, sheet_name='Resultados')
            
            # Ajustar largura da coluna de conversa para não truncar
            worksheet = writer.sheets['Resultados']
            if "conversa" in df_download_completo.columns:
                col_idx = df_download_completo.columns.get_loc("conversa") + 1
                worksheet.column_dimensions[worksheet.cell(row=1, column=col_idx).column_letter].width = 100
                # Habilitar quebra de texto
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if Alignment:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        excel_data = excel_buffer.getvalue()
        
        st.download_button(
            label="📥 Download Excel (Completo)",
            data=excel_data,
            file_name=f"relatorio_qa_completo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )


